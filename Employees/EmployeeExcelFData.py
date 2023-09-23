import pandas as pd
from datetime import datetime
from openpyxl import Workbook


class EmployeeExcelData:

    def __init__(self, input_csv, output_xlsx):
        self.input_csv = input_csv  # Ім'я вхідного CSV-файлу
        self.output_xlsx = output_xlsx  # Ім'я вихідного XLSX-файлу

    def age_calculator(self, dob):
        """Обчислення віку на основі дати народження."""
        today = datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age

    def generate_excel(self):
        try:
            # Завантаження даних із CSV-файлу у DataFrame
            data_frame = pd.read_csv(self.input_csv, encoding='utf-8-sig', sep=';', parse_dates=['Дата народження'], dayfirst=True)

            # Обчислення віку
            data_frame['Вік'] = data_frame['Дата народження'].apply(self.age_calculator)

            # Видалення непотрібних стовпців
            data_frame.drop(columns=['Стать', 'Посада', 'Місто проживання', 'Адреса проживання', 'Телефон', 'Email'], inplace=True)

            # Створення нового робочого зошита Excel
            workbook = Workbook()
            all_sheet = workbook.active
            all_sheet.title = "all"

            # Додаємо заголовки на аркуш 'all'
            headers = ["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"]
            for col_num, header in enumerate(headers, 1):
                all_sheet.cell(row=1, column=col_num, value=header)

            # Записуємо дані на аркуш 'all'
            for row_index, row_data in enumerate(data_frame.iterrows(), 2):  # Починаємо з 2 рядка, оскільки 1 рядок - заголовки
                all_sheet.cell(row=row_index, column=1, value=row_index-1)  # Додаємо номер рядка
                for col_index, value in enumerate(row_data[1], 2):  # Починаємо з 2 стовпця, оскільки 1 стовпець - номер рядка
                    all_sheet.cell(row=row_index, column=col_index, value=value)

            # Додаємо дані на інші аркуші відповідно до категорій віку
            age_sheets = {
                "younger_18": (0, 17),
                "18-45": (18, 45),
                "45-70": (46, 70),
                "older_70": (71, 200)
            }

            for sheet_title, age_range in age_sheets.items():
                sheet = workbook.create_sheet(title=sheet_title)
                filtered_data = data_frame[data_frame['Вік'].between(*age_range)]

                # Додаємо заголовки на кожний аркуш
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)

                # Записуємо дані на кожний аркуш
                for row_index, row_data in enumerate(filtered_data.iterrows(), 2):
                    sheet.cell(row=row_index, column=1, value=row_index-1)  # Додаємо номер рядка
                    for col_index, value in enumerate(row_data[1], 2):
                        sheet.cell(row=row_index, column=col_index, value=value)

            # Збереження файлу Excel
            workbook.save(self.output_xlsx)
            print("Ok. Файл успішно згенеровано")

        except pd.errors.ParserError:
            print(f"Помилка: Неможливо проаналізувати CSV-файл: {self.input_csv}")

        except PermissionError:
            print(f"Помилка: Неможливо створити файл XLSX: {self.output_xlsx} (Доступ заборонено)")

        except Exception as e:
            print(f"Помилка: {e}")


# Використання:
excel_generator = EmployeeExcelData(
    input_csv="employees.csv",
    output_xlsx="employees.xlsx"
)
excel_generator.generate_excel()
